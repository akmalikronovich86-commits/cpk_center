import os

# Создаём директории
os.makedirs('/Users/pro16/cpk_center/templates/admin/users/studentprofile', exist_ok=True)

# Файл 1: admin.py
admin_py = '''from django import forms
from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import Group
from django.contrib.auth.admin import GroupAdmin
from django_jsonform.widgets import JSONFormWidget
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import path
from django.shortcuts import render
from django.db import transaction
import openpyxl
from openpyxl import Workbook
from apps.users.models import User, StudentProfile, LecturerProfile, Module
from datetime import datetime


class ModuleInline(admin.TabularInline):
    model = Module
    extra = 1
    fields = ['name', 'hours', 'description', 'is_active']
    show_change_link = True
    verbose_name = "Modul"
    verbose_name_plural = "Modullar"


class LecturerProfileForm(forms.ModelForm):
    class Meta:
        model = LecturerProfile
        fields = '__all__'
        widgets = {
            'working_hours': JSONFormWidget(
                schema={
                    'type': 'object',
                    'properties': {
                        'dushanba': {'type': 'array', 'items': {'type': 'string'}},
                        'seshanba': {'type': 'array', 'items': {'type': 'string'}},
                        'chorshanba': {'type': 'array', 'items': {'type': 'string'}},
                        'payshanba': {'type': 'array', 'items': {'type': 'string'}},
                        'juma': {'type': 'array', 'items': {'type': 'string'}},
                        'shanba': {'type': 'array', 'items': {'type': 'string'}},
                        'yakshanba': {'type': 'array', 'items': {'type': 'string'}},
                    }
                }
            )
        }


class StudentProfileForm(forms.ModelForm):
    class Meta:
        model = StudentProfile
        fields = '__all__'
        widgets = {
            'birth_date': forms.TextInput(attrs={'placeholder': 'MM/DD/YY'}),
            'qualification_period': forms.TextInput(attrs={'placeholder': 'DD.MM-DD.MM.YYYY'}),
        }


@admin.register(StudentProfile)
class StudentProfileAdmin(admin.ModelAdmin):
    form = StudentProfileForm
    list_display = (
        'sequence_number', 'full_name_qualification', 'position', 
        'group', 'regional_branch', 'final_grade', 'qualification_period'
    )
    search_fields = ('full_name_qualification', 'group', 'xtm', 'passport')
    list_editable = ('final_grade',)
    actions = ['export_to_excel']
    
    fieldsets = (
        ("Asosiy ma'lumotlar", {
            'fields': ('user', 'sequence_number', 'xtm', 'full_name_qualification')
        }),
        ("Shaxsiy ma'lumotlar", {
            'fields': ('position', 'passport', 'birth_date', 'phone_number')
        }),
        ("O'quv ma'lumotlari", {
            'fields': ('group', 'regional_branch', 'district_power_supply')
        }),
        ('Natijalar', {
            'fields': ('final_grade', 'independent_study_topic', 'qualification_period')
        }),
        ('Tizim ma\'lumotlari', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ('created_at', 'updated_at')
    
    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tinglovchilar"
        
        headers = [
            'T/R', 'XTM', "Malaka oshiruvchilar", 'Lavozimi', 
            'Pasport', "Tug'ilgan sana", 'Guruhi', 'Hududiy filiali nomi',
            'Tuman elektr ta\'minoti', 'Telefon raqami', 'Yakuniy bahosi',
            "Mustaqil ta'lim mavzusi", 'Malaka oshirish muddati'
        ]
        ws.append(headers)
        
        for profile in queryset:
            ws.append([
                profile.sequence_number or '',
                profile.xtm or '',
                profile.full_name_qualification or '',
                profile.position or '',
                profile.passport or '',
                profile.birth_date or '',
                profile.group or '',
                profile.regional_branch or '',
                profile.district_power_supply or '',
                profile.phone_number or '',
                profile.final_grade or '',
                profile.independent_study_topic or '',
                profile.qualification_period or '',
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Tinglovchilar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        
        self.message_user(request, f"{queryset.count()} ta tinglovchi eksport qilindi")
        return response
    
    export_to_excel.short_description = "Tanlangan tinglovchilarni Excel'ga eksport qilish"
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-excel/',
                self.admin_site.admin_view(self.import_excel_view),
                name='users_studentprofile_import_excel'
            ),
        ]
        return custom_urls + urls
    
    def import_excel_view(self, request):
        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            
            if not excel_file:
                messages.error(request, "Fayl yuklanmagan!")
                return HttpResponseRedirect(request.path)
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "Faqat .xlsx yoki .xls formatidagi fayllar qabul qilinadi!")
                return HttpResponseRedirect(request.path)
            
            try:
                with transaction.atomic():
                    result = self.process_excel_import(excel_file)
                    
                    if result['errors']:
                        error_msg = f"{result['created']} ta yangi, {result['updated']} ta yangilandi. "
                        error_msg += f"{len(result['errors'])} ta qatorda xatolik: "
                        error_msg += "; ".join(result['errors'][:5])
                        if len(result['errors']) > 5:
                            error_msg += f" ... va yana {len(result['errors']) - 5} ta"
                        messages.warning(request, error_msg)
                    else:
                        messages.success(
                            request, 
                            f"Muvaffaqiyatli! {result['created']} ta yangi, "
                            f"{result['updated']} ta yangilandi. Jami: {result['total']} ta tinglovchi."
                        )
                    
                    return HttpResponseRedirect(
                        f'/admin/users/studentprofile/?created={result["created"]}&updated={result["updated"]}'
                    )
                    
            except Exception as e:
                messages.error(request, f"Xatolik yuz berdi: {str(e)}")
                return HttpResponseRedirect(request.path)
        
        return render(request, 'admin/users/studentprofile/import_excel.html', {
            'title': 'Excel dan import qilish',
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        })
    
    def process_excel_import(self, excel_file):
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        result = {
            'created': 0,
            'updated': 0,
            'total': 0,
            'errors': [],
        }
        
        header_row_idx = 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            row_values = [str(cell).strip() if cell else '' for cell in row]
            if any(keyword in ' '.join(row_values) for keyword in ['Malaka', 'T/R', 'XTM']):
                header_row_idx = row_idx
                break
        
        header_row = [str(cell).strip().lower() if cell else '' for cell in list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]]
        
        column_map = self._build_column_map(header_row)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            if not any(row):
                continue
            
            try:
                data = self._extract_row_data(row, column_map)
                
                if not data.get('full_name_qualification'):
                    continue
                
                result['total'] += 1
                
                profile, created = self._create_or_update_profile(data)
                
                if created:
                    result['created'] += 1
                else:
                    result['updated'] += 1
                    
            except Exception as e:
                result['errors'].append(f"Qator {row_idx}: {str(e)}")
                continue
        
        return result
    
    def _build_column_map(self, header_row):
        column_map = {}
        keywords = {
            'sequence_number': ['t/r', 'tr', 'tartib'],
            'xtm': ['xtm'],
            'full_name_qualification': ['malaka', 'oshiruvchilar', 'f.i.sh', 'fio'],
            'position': ['lavozimi', 'lavozim'],
            'passport': ['pasport'],
            'birth_date': ['tug', 'ilgan', 'sana'],
            'group': ['guruh'],
            'regional_branch': ['hududiy', 'filial'],
            'district_power_supply': ['tuman', 'elektr', 'ta'],
            'phone_number': ['telefon'],
            'final_grade': ['yakuniy', 'baho'],
            'independent_study_topic': ['mustaqil', 'ta', 'lim', 'mavzu'],
            'qualification_period': ['malaka', 'muddati', 'davri'],
        }
        
        for col_idx, header in enumerate(header_row):
            header_lower = header.lower().strip()
            for field, kws in keywords.items():
                if any(kw in header_lower for kw in kws) and field not in column_map:
                    column_map[field] = col_idx
                    break
        
        return column_map
    
    def _extract_row_data(self, row, column_map):
        data = {}
        
        def get_value(field):
            if field in column_map and column_map[field] < len(row):
                value = row[column_map[field]]
                return str(value).strip() if value is not None else ''
            return ''
        
        data['sequence_number'] = get_value('sequence_number')
        data['xtm'] = get_value('xtm')
        data['full_name_qualification'] = get_value('full_name_qualification')
        data['position'] = get_value('position')
        data['passport'] = get_value('passport')
        data['birth_date'] = self.format_date_value(get_value('birth_date'), row[column_map.get('birth_date', 0)] if 'birth_date' in column_map else None)
        data['group'] = get_value('group')
        data['regional_branch'] = get_value('regional_branch')
        data['district_power_supply'] = get_value('district_power_supply')
        data['phone_number'] = get_value('phone_number')
        data['final_grade'] = get_value('final_grade')
        data['independent_study_topic'] = get_value('independent_study_topic')
        data['qualification_period'] = get_value('qualification_period')
        
        return data
    
    def format_date_value(self, date_str, raw_value):
        if not date_str or date_str == 'None':
            return ''
        
        if isinstance(raw_value, datetime):
            return raw_value.strftime('%m/%d/%y')
        
        if isinstance(raw_value, (int, float)):
            try:
                from datetime import timedelta
                base_date = datetime(1899, 12, 30)
                return (base_date + timedelta(days=int(raw_value))).strftime('%m/%d/%y')
            except:
                pass
        
        return str(date_str)
    
    def _create_or_update_profile(self, data):
        full_name = data.get('full_name_qualification', '').strip()
        
        if not full_name:
            raise ValueError("F.I.Sh. bo'sh")
        
        passport = data.get('passport', '').strip().replace(' ', '')
        if passport:
            username = f"student_{passport}"
        else:
            import hashlib
            hash_suffix = hashlib.md5(full_name.encode()).hexdigest()[:8]
            username = f"student_{hash_suffix}"
        
        username = username[:150]
        
        user = None
        
        if passport:
            try:
                user = User.objects.get(username=username)
            except User.DoesNotExist:
                user = User.objects.create_user(
                    username=username,
                    password=User.objects.make_random_password(),
                    role='student',
                    full_name=full_name,
                )
        
        if not user:
            user = User.objects.create_user(
                username=username,
                password=User.objects.make_random_password(),
                role='student',
                full_name=full_name,
            )
        
        profile, created = StudentProfile.objects.update_or_create(
            user=user,
            defaults={
                'sequence_number': data.get('sequence_number', '')[:20],
                'xtm': data.get('xtm', '')[:50],
                'full_name_qualification': full_name[:200],
                'position': data.get('position', '')[:100],
                'passport': passport[:50],
                'birth_date': data.get('birth_date', '')[:50],
                'group': data.get('group', '')[:50],
                'regional_branch': data.get('regional_branch', '')[:100],
                'district_power_supply': data.get('district_power_supply', '')[:100],
                'phone_number': data.get('phone_number', '')[:50],
                'final_grade': data.get('final_grade', '')[:50],
                'independent_study_topic': data.get('independent_study_topic', ''),
                'qualification_period': data.get('qualification_period', '')[:50],
            }
        )
        
        return profile, created
    
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(LecturerProfile)
class LecturerProfileAdmin(admin.ModelAdmin):
    form = LecturerProfileForm
    inlines = [ModuleInline]
    list_display = ('user', 'specialization', 'experience_years')
    search_fields = ('user__username', 'user__first_name', 'user__last_name')


@admin.register(Module)
class ModuleAdmin(admin.ModelAdmin):
    list_display = ('name', 'lecturer', 'hours', 'is_active', 'created_at')
    list_filter = ('is_active', 'created_at')
    search_fields = ('name', 'lecturer__user__username')
    list_editable = ('is_active',)


admin.site.register(User, UserAdmin)

try:
    admin.site.unregister(Group)
except admin.sites.NotRegistered:
    pass

@admin.register(Group)
class CustomGroupAdmin(GroupAdmin):
    pass

Group._meta.verbose_name = 'Guruh'
Group._meta.verbose_name_plural = 'Guruhlar'
'''

with open('/Users/pro16/cpk_center/apps/users/admin.py', 'w', encoding='utf-8') as f:
    f.write(admin_py)

print("✅ admin.py yaratildi!")

# Файл 2: import_excel.html
import_html = '''{% extends "admin/base_site.html" %}
{% load i18n static %}

{% block extrastyle %}
{{ block.super }}
<style>
    .import-container {
        max-width: 800px;
        margin: 20px auto;
        padding: 20px;
        background: #fff;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .import-container h1 {
        color: #417690;
        margin-bottom: 20px;
        border-bottom: 2px solid #417690;
        padding-bottom: 10px;
    }
    .import-form {
        margin: 20px 0;
        padding: 20px;
        background: #f8f8f8;
        border-radius: 5px;
    }
    .import-form label {
        display: block;
        font-weight: bold;
        margin-bottom: 10px;
        color: #333;
    }
    .import-form input[type="file"] {
        margin: 10px 0;
        padding: 10px;
        border: 2px dashed #417690;
        background: #fff;
        width: 100%;
        cursor: pointer;
    }
    .import-form button {
        background: #417690;
        color: white;
        padding: 10px 30px;
        border: none;
        border-radius: 4px;
        cursor: pointer;
        font-size: 14px;
        margin-top: 10px;
    }
    .import-form button:hover {
        background: #205067;
    }
    .info-box {
        background: #e7f3ff;
        border-left: 4px solid #2196F3;
        padding: 15px;
        margin: 20px 0;
        border-radius: 4px;
    }
    .info-box h3 {
        color: #2196F3;
        margin-top: 0;
    }
    .column-list {
        background: #fff;
        padding: 15px;
        border-radius: 4px;
        margin: 10px 0;
    }
    .column-list ol {
        margin: 0;
        padding-left: 20px;
    }
    .column-list li {
        padding: 5px 0;
        color: #555;
    }
    .warning-box {
        background: #fff3cd;
        border-left: 4px solid #ffc107;
        padding: 15px;
        margin: 20px 0;
        border-radius: 4px;
    }
    .success-box {
        background: #d4edda;
        border-left: 4px solid #28a745;
        padding: 15px;
        margin: 20px 0;
        border-radius: 4px;
    }
    .error-box {
        background: #f8d7da;
        border-left: 4px solid #dc3545;
        padding: 15px;
        margin: 20px 0;
        border-radius: 4px;
    }
    .back-link {
        display: inline-block;
        margin-top: 20px;
        color: #417690;
        text-decoration: none;
    }
    .back-link:hover {
        text-decoration: underline;
    }
</style>
{% endblock %}

{% block content %}
<div class="import-container">
    <h1>Excel fayldan Tinglovchilarni import qilish</h1>

    {% if messages %}
        {% for message in messages %}
            <div class="{% if message.tags == 'success' %}success-box{% elif message.tags == 'error' %}error-box{% else %}info-box{% endif %}">
                {{ message }}
            </div>
        {% endfor %}
    {% endif %}

    <div class="info-box">
        <h3>Faylga qo'yiladigan talablar:</h3>
        <p>Excel fayl (.xlsx) quyidagi ustunlarga ega bo'lishi kerak:</p>
        <div class="column-list">
            <ol>
                <li><strong>T/R</strong> - Tartib raqami</li>
                <li><strong>XTM</strong> - XTM kodi</li>
                <li><strong>Malaka oshiruvchilar</strong> - F.I.Sh. (to'liq ism)</li>
                <li><strong>Lavozimi</strong> - Egallab turgan lavozimi</li>
                <li><strong>Pasport</strong> - Pasport seriya va raqami</li>
                <li><strong>Tug'ilgan sana</strong> - Tug'ilgan sanasi</li>
                <li><strong>Guruhi</strong> - O'quv guruhi (masalan: 26-01 guruh)</li>
                <li><strong>Hududiy filiali nomi</strong> - Hududiy filial</li>
                <li><strong>Tuman elektr ta'minoti</strong> - Tuman E'T</li>
                <li><strong>Telefon raqami</strong> - Aloqa raqami</li>
                <li><strong>Yakuniy bahosi</strong> - Yakuniy baho</li>
                <li><strong>Mustaqil ta'lim mavzusi</strong> - Mavzu nomi</li>
                <li><strong>Malaka oshirish muddati</strong> - Davr (masalan: 05.01-16.01.2026)</li>
            </ol>
        </div>
    </div>

    <div class="warning-box">
        <h3>Muhim eslatma:</h3>
        <ul>
            <li>Fayl <strong>.xlsx</strong> formatida bo'lishi kerak</li>
            <li>Birinchi qator - sarlavhalar (ustun nomlari)</li>
            <li>Ikkinchi qatordan boshlab - ma'lumotlar</li>
            <li>Bo'sh qatorlar avtomatik o'tkazib yuboriladi</li>
            <li>Agar F.I.Sh. mavjud bo'lsa, yangilash amalga oshiriladi (takrorlanmaydi)</li>
        </ul>
    </div>

    <div class="import-form">
        <form method="post" enctype="multipart/form-data">
            {% csrf_token %}
            <label for="excel_file">Excel faylni tanlang:</label>
            <input type="file" name="excel_file" id="excel_file" accept=".xlsx,.xls" required>
            <br>
            <button type="submit">Import qilish</button>
        </form>
    </div>

    <a href="{% url 'admin:users_studentprofile_changelist' %}" class="back-link">
        ← Tinglovchilar ro'yxatiga qaytish
    </a>
</div>
{% endblock %}
'''

with open('/Users/pro16/cpk_center/templates/admin/users/studentprofile/import_excel.html', 'w', encoding='utf-8') as f:
    f.write(import_html)

print("✅ import_excel.html yaratildi!")

# Файл 3: change_list.html
change_list_html = '''{% extends "admin/change_list.html" %}
{% load i18n %}

{% block object-tools-items %}
    {% if show_import_button %}
    <li>
        <a href="{% url 'admin:users_studentprofile_import_excel' %}" 
           style="background: #28a745; color: white; padding: 8px 15px; border-radius: 4px; text-decoration: none; font-weight: bold;">
            Excel dan import
        </a>
    </li>
    {% endif %}
    {{ block.super }}
{% endblock %}
'''

with open('/Users/pro16/cpk_center/templates/admin/users/studentprofile/change_list.html', 'w', encoding='utf-8') as f:
    f.write(change_list_html)

print("✅ change_list.html yaratildi!")
print("\n🎉 Barcha fayllar muvaffaqiyatli yaratildi!")
