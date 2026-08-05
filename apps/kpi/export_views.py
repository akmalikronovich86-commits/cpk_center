import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from apps.courses.models import Course
from apps.schedules.models import Attendance
from apps.users.models import User
from datetime import datetime, timedelta
from django.utils import timezone


@staff_member_required
def export_kpi_to_excel(request):
    """Экспорт KPI данных в Excel"""
    # Создаем workbook и worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Hisoboti"
    
    # Стили для заголовков
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заголовок отчета
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "KPI Hisoboti"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Дата отчета
    ws.merge_cells('A2:D2')
    date_cell = ws['A2']
    date_cell.value = f"Sana: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
    date_cell.alignment = Alignment(horizontal="center")
    
    # === Секция 1: Курсы по лекторам ===
    row = 4
    ws.cell(row=row, column=1, value="O'qituvchilar bo'yicha kurslar soni").font = Font(bold=True, size=12)
    row += 1
    
    # Заголовки таблицы
    headers = ["O'qituvchi", "Kurslar soni"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Данные по лекторам
    current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    current_month_end = current_month_start + timedelta(days=32)
    current_month_end = current_month_end.replace(day=1)
    
    courses_by_lecturer = Course.objects.filter(
        updated_at__gte=current_month_start,
        updated_at__lte=current_month_end
    ).values('lecturer__last_name', 'lecturer__first_name').annotate(count=len('id')).order_by('-count')
    
    # Если annotate не работает, используем альтернативный подход
    from django.db.models import Count
    courses_by_lecturer = Course.objects.filter(
        updated_at__gte=current_month_start,
        updated_at__lte=current_month_end
    ).values('lecturer__last_name', 'lecturer__first_name').annotate(count=Count('id')).order_by('-count')
    
    row += 1
    for item in courses_by_lecturer:
        lecturer_name = f"{item.get('lecturer__last_name', '')} {item.get('lecturer__first_name', '')}".strip()
        ws.cell(row=row, column=1, value=lecturer_name)
        ws.cell(row=row, column=2, value=item['count'])
        row += 1
    
    # === Секция 2: Посещаемость ===
    row += 2
    ws.cell(row=row, column=1, value="Davomat statistikasi").font = Font(bold=True, size=12)
    row += 1
    
    # Заголовки
    headers = ["Holat", "Soni"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Данные по посещаемости
    attendance_stats = Attendance.objects.values('status').annotate(count=Count('id')).order_by('-count')
    
    row += 1
    for item in attendance_stats:
        ws.cell(row=row, column=1, value=item['status'])
        ws.cell(row=row, column=2, value=item['count'])
        row += 1
    
    # === Секция 3: Общая статистика ===
    row += 2
    ws.cell(row=row, column=1, value="Umumiy statistika").font = Font(bold=True, size=12)
    row += 1
    
    total_courses = Course.objects.count()
    total_attendance = Attendance.objects.count()
    total_students = User.objects.filter(role='student').count()
    
    stats = [
        ("Jami kurslar", total_courses),
        ("Jami davomat yozuvlari", total_attendance),
        ("Jami o'quvchilar", total_students),
    ]
    
    for label, value in stats:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    # Создаем HTTP response с Excel файлом
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=KPI_Report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response
