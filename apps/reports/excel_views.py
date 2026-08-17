"""
Bir tugma bilan Excel hisobotlar - openpyxl
"""
from datetime import datetime

from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.core.decorators import role_required
from apps.certificates.models import Certificate
from apps.groups.models import StudentRecord
from apps.schedules.models import Schedule
from apps.users.models import LecturerProfile

STAFF_ROLES = ('admin', 'director', 'methodist', 'department_head')


def _xlsx_response(headers, rows, sheet_name, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)
    fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append(list(row))

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 24
    ws.freeze_panes = 'A2'

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@role_required(*STAFF_ROLES)
def reports_index(request):
    """Hisobotlar sahifasi"""
    return render(request, 'reports/index.html')


@role_required(*STAFF_ROLES)
def export_students(request):
    headers = ['ID', 'F.I.Sh.', 'Guruh', 'Telefon', 'Lavozimi',
               'Hududiy filial', 'Yakuniy baho', 'Malaka muddati']
    rows = []
    for s in StudentRecord.objects.order_by('id'):
        rows.append((
            s.id,
            s.full_name or '',
            s.group or '',
            s.phone or '',
            getattr(s, 'position', '') or '',
            getattr(s, 'regional_branch', '') or '',
            getattr(s, 'final_grade', '') or '',
            getattr(s, 'qualification_period', '') or '',
        ))
    return _xlsx_response(headers, rows, 'Tinglovchilar', f'tinglovchilar_{datetime.now():%Y%m%d}.xlsx')


@role_required(*STAFF_ROLES)
def export_certificates(request):
    headers = ['Sertifikat raqami', 'Seriya', 'Berilgan sana', 'Amal qilish muddati',
               'Holat', 'MYHT reyestri', 'QR kod']
    rows = [(
        c.certificate_number or '', c.series or '',
        c.issue_date.strftime('%d.%m.%Y') if c.issue_date else '',
        c.expiry_date.strftime('%d.%m.%Y') if c.expiry_date else '',
        c.get_status_display(), c.registry_number or '', c.qr_code or '',
    ) for c in Certificate.objects.order_by('-issue_date')]
    return _xlsx_response(headers, rows, 'Sertifikatlar', f'sertifikatlar_{datetime.now():%Y%m%d}.xlsx')


@role_required(*STAFF_ROLES)
def export_schedule(request):
    """Jadval eksporti - PRAVILNYE POLYA"""
    headers = ['Boshlanish', 'Tugash', 'Guruh', 'Mavzu', "Ma'ruzachi", 'Xona', 'Holat']
    rows = []
    for sch in Schedule.objects.order_by('date_start'):
        rows.append((
            sch.date_start.strftime('%d.%m.%Y %H:%M') if sch.date_start else '',
            sch.date_end.strftime('%d.%m.%Y %H:%M') if sch.date_end else '',
            str(getattr(sch, 'group', '') or ''),
            str(getattr(sch, 'topic', '') or ''),
            getattr(getattr(getattr(sch, 'lecturer', None), 'user', None), 'full_name', '') or '',
            getattr(sch, 'room', '') or '',
            getattr(sch, 'status', '') or '',
        ))
    return _xlsx_response(headers, rows, 'Jadval', f'jadval_{datetime.now():%Y%m%d}.xlsx')


@role_required(*STAFF_ROLES)
def export_lecturers(request):
    headers = ['F.I.Sh.', 'Mutaxassislik', 'Tajriba (yil)', 'Telefon']
    rows = [(
        getattr(l.user, 'full_name', '') or l.user.username,
        l.specialization or '', l.experience_years,
        getattr(l.user, 'phone', '') or '',
    ) for l in LecturerProfile.objects.select_related('user')]
    return _xlsx_response(headers, rows, "Ma'ruzachilar", f'maruzachilar_{datetime.now():%Y%m%d}.xlsx')
